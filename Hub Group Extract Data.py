import PyPDF2
import re
import ctypes
from ctypes import wintypes
import csv
from datetime import datetime
import subprocess
from tkinter import messagebox
import keyboard
import win32gui
import win32con
import os
import time
import openpyxl
import pyautogui
import psutil
import pygetwindow as gw

# Specify the folder containing the PDF files
pdf_folder_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Invoices"

# Get a list of all PDF files in the folder
pdf_files = [f for f in os.listdir(pdf_folder_path) if f.lower().endswith('.pdf')]

def is_excel_file_open(file_path):
    for process in psutil.process_iter():
        try:
            if process.name() == "EXCEL.EXE":
                for file in process.open_files():
                    if file_path.lower() in file.path.lower():
                        return True
        except Exception:
            pass
    return False

def check_for_rjw(text):
    if "RJW" in text:
        return "RJW Logistics W10"
    else:
        return "NOT RJW, PLEASE CHECK"

def view_hub_group_invoices():
    import openpyxl
    import pygetwindow as gw
    from time import sleep
    import fnmatch
    import tkinter as tk
    import os
    import subprocess
    import platform
    from tkinter import messagebox

    def get_unique_invoice_numbers(sheet):
        invoice_numbers = set()
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            invoice_number = row[0]
            if invoice_number:
                invoice_numbers.add(invoice_number)
        return list(invoice_numbers)

    def show_loading_popup():
        popup = tk.Tk()
        popup.wm_title("Loading")
        label = tk.Label(popup, text="Loading the next PDF, please wait...")
        label.pack(side="top", fill="x", pady=10)
        popup.after(750, popup.destroy)  # Auto-close the popup after 3 seconds
        popup.mainloop()

    def prompt_order_number(invoice_number):
        root = tk.Tk()
        root.wm_title("Order Number")  # Set the window title

        # Make the window always on top
        root.attributes('-topmost', True)

        # Specify the position of the window (x=200, y=200)
        root.geometry("+400+400")

        tk.Label(root, text=f"Enter the order number for invoice {invoice_number}:").grid(row=0)
        entry = tk.Entry(root)
        entry.grid(row=1)
        order_number = None

        def on_button_click(event=None):
            nonlocal order_number
            order_number = entry.get()
            root.destroy()

        submit_button = tk.Button(root, text="Submit", command=on_button_click)
        submit_button.grid(row=2)

        # Bind the Enter key to the submit function
        root.bind('<Return>', on_button_click)

        # Update idle tasks, lift and focus the window
        root.update_idletasks()
        root.lift()
        root.focus_force()

        entry.focus_set()  # Set focus to the entry widget
        root.mainloop()
        return order_number

    # Load the existing workbook
    excel_file_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Hub Group Upload.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the "Data" sheet
    sheet_name = 'Data'
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        print("Sheet not found")
        exit()

    # Get unique invoice numbers
    unique_invoice_numbers = get_unique_invoice_numbers(sheet)

    # Specify the folder containing the PDF files
    pdf_folder_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Invoices"

    # Set NitroPDF as the PDF viewer
    pdf_viewer = r"C:\Program Files\Nitro\Pro\13\NitroPDF.exe"

    # Initialize a dictionary to store the order numbers associated with each invoice number
    order_numbers = {}

    previous_pdf_process = None
    first_iteration = True

    for invoice_number in unique_invoice_numbers:
        pdf_file_name = f"{invoice_number}_Invoice.pdf"
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file_name)

        if os.path.exists(pdf_file_path):
            # Close the previous PDF if it exists
            # Commented out the following lines to keep the previous PDF open
            # if previous_pdf_process is not None:
            #     previous_pdf_process.kill()

            pdf_process = subprocess.Popen([pdf_viewer, pdf_file_path])

            # Add extra sleep time for the first iteration to let NitroPDF fully load
            if first_iteration:
                sleep(3)  # Adjust this sleep time as needed
                first_iteration = False
            else:
                sleep(1)

            pdf_window = None
            for win in gw.getAllWindows():
                if fnmatch.fnmatch(win.title, f"{pdf_file_name} - *"):
                    pdf_window = win
                    break

            if pdf_window is None:
                print(f"Unable to find the NitroPDF window for invoice number: {invoice_number}")
                pdf_process.kill()
                continue
            pdf_window.activate()

            # Prompt the user to enter the order number and save it in the dictionary
            order_number = prompt_order_number(invoice_number)
            order_numbers[invoice_number] = order_number

            previous_pdf_process = pdf_process  # Save the current process as the previous one
        else:
            print(f"PDF file not found for invoice number: {invoice_number}")

    print("All PDFs have been viewed.")

    # All PDFs have been viewed, now update the Excel workbook with the order numbers
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=False):
        invoice_number = row[0].value
        if invoice_number in order_numbers:
            order_number = order_numbers[invoice_number]
            sheet.cell(row=row[0].row, column=8).value = order_number

    # Save and close the workbook
    workbook.save(excel_file_path)
    workbook.close()

    # Close NitroPDF processes
    if previous_pdf_process is not None:
        previous_pdf_process.kill()

    # Find any remaining NitroPDF processes and kill them
    system = platform.system()
    if system == 'Windows':
        subprocess.run(['taskkill', '/IM', 'NitroPDF.exe', '/F'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    elif system == 'Linux' or system == 'Darwin':
        subprocess.run(['pkill', '-f', 'NitroPDF'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    # Show a message box indicating that all order numbers have been imported
    root = tk.Tk()
    root.withdraw()  # Hide the main tkinter window
    messagebox.showinfo("Order Numbers Imported", "All order numbers have been imported.")

def extract_invoice_data_from_pdf(pdf_file_path):
    with open(pdf_file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        first_page_text = reader.pages[0].extract_text()

    invoice_number, reference_numbers, invoice_date, num_po_numbers = extract_invoice_data(first_page_text)
    return invoice_number, reference_numbers, invoice_date, first_page_text

def modify_charge_type(charge_type):
    if charge_type == "Line Haul":
        return ""
    else:
        return "- " + charge_type

def wait_for_key_combination2():
    CTRL_PRESSED = 0x8000
    S_KEY = 0x53

    while True:
        ctrl_key_state = ctypes.windll.user32.GetAsyncKeyState(wintypes.VK_CONTROL)
        s_key_state = ctypes.windll.user32.GetAsyncKeyState(S_KEY)

        if (ctrl_key_state & CTRL_PRESSED) and (s_key_state & 0x8000):
            break

def extract_po_number(reference_line_text):
    po_numbers = re.findall(r'\b(PO)?(\d{5}[A-Za-z]?)(?=\b|\s*PIECES|,)', reference_line_text)
    po_numbers = ['PO' + num[1] for num in po_numbers]
    return po_numbers

def extract_invoice_data(text):
    invoice_number = re.search(r'INVOICE#\s+(\d+)', text)
    reference_line = re.search(r'REFERENCE 1#(.*?PIECES)', text, re.DOTALL)  # Updated regex pattern
    invoice_date = re.search(r'INVOICE DATE\s+(\d{2}/\d{2}/\d{4})', text)

    if invoice_number:
        invoice_number = invoice_number.group(1)
    else:
        invoice_number = None

    if reference_line:
        reference_line_text = reference_line.group(1)
        print(f"Reference Line: {reference_line_text}")  # Print the extracted reference line
        po_numbers = extract_po_number(reference_line_text)
        reference_numbers = po_numbers  # Use po_numbers directly without adding 'PO' prefix again
    else:
        reference_numbers = []

    if reference_numbers:
        num_po_numbers = len(reference_numbers)
    else:
        num_po_numbers = 1  # Set to 1 if the list is empty to avoid division by zero

    if invoice_date:
        invoice_date = invoice_date.group(1)
    else:
        invoice_date = None

    return invoice_number, reference_numbers, invoice_date, num_po_numbers

def extract_charges_from_pdf(pdf_file_path):
    with open(pdf_file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        first_page_text = reader.pages[0].extract_text()

    start_index = first_page_text.index("CHARGE DESCRIPTION              AMOUNT DUE CURRENCY")
    end_index = first_page_text.index("BILL MEMO")
    charges_text = first_page_text[start_index:end_index].strip()

    lines = charges_text.split('\n')[1:]
    charge_description = ""
    valid_matches = []

    for line in lines:
        if line.strip() == "":
            continue

        amount_match = re.search(r'(\d{1,3}(?:,\d{3})*\.\d{2})', line)
        if "USD" in line or amount_match:
            if amount_match:
                amount = amount_match.group(1)

            if not charge_description:
                charge_description = re.sub(r'\s?\d{1,3}(?:,\d{3})*\.\d{2}.*', '', line).strip()

            modified_charge_description = re.sub(r'^Transload,', 'Transload, Warehouse Charge',
                                                 charge_description.strip())
            valid_matches.append((modified_charge_description, amount.replace(",", "")))
            charge_description = ""
        else:
            charge_description += " " + line.strip()

    return valid_matches, first_page_text

def wait_for_key_combination():
    import keyboard

    event = keyboard.read_event()
    while not (event.name == '`' and event.event_type == keyboard.KEY_DOWN and keyboard.is_pressed('ctrl')):
        event = keyboard.read_event()

def clear_sheet(sheet):
    sheet.delete_rows(1, sheet.max_row)
    sheet.delete_cols(1, sheet.max_column)

def clear_worksheet_contents(worksheet, start_row=1):
    for row in range(start_row, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).value = None

def copy_column(source_sheet, source_column, target_sheet, target_column):
    for row in range(1, source_sheet.max_row + 1):
        target_sheet.cell(row=row, column=target_column).value = source_sheet.cell(row=row, column=source_column).value

def remove_duplicates(sheet, col_index, row_offset=1):
    unique_values = set()
    row_numbers_to_delete = []

    for row in sheet.iter_rows(min_row=row_offset + 1):
        cell_value = row[col_index].value
        if cell_value not in unique_values:
            unique_values.add(cell_value)
        else:
            row_numbers_to_delete.append(row[col_index].row)

    for row_number in reversed(row_numbers_to_delete):
        sheet.delete_rows(row_number)

    return sheet

def copy_columns(source_sheet, destination_sheet, start_col, end_col):
    for row_num, row in enumerate(source_sheet.iter_rows(min_row=1, min_col=start_col, max_col=end_col), start=1):
        for col_num, cell in enumerate(row, start=start_col):
            destination_sheet.cell(row=row_num, column=col_num, value=cell.value)

def process_worksheet(workbook, worksheet_name, date_folder_path):
    # Select the worksheet by name
    source_worksheet = workbook[worksheet_name]

    # Create a new workbook
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    # Copy values from source worksheet to new worksheet
    for row in source_worksheet.iter_rows(values_only=True):
        new_worksheet.append(row)

    # Set the file name format based on the worksheet_name
    if worksheet_name == 'Hub Group Primary':
        csv_file_name = f"Primary - {current_date}.csv"
    elif worksheet_name == 'Hub Group Items':
        csv_file_name = f"Items - {current_date}.csv"
    else:
        csv_file_name = f"{worksheet_name} - {current_date}.csv"

    csv_file_path = os.path.join(date_folder_path, csv_file_name)

    # Save the new worksheet as CSV
    save_worksheet_as_csv(new_worksheet, csv_file_path)

def save_worksheet_as_csv(worksheet, csv_file_path):
    with open(csv_file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(row)

def show_message_box(title, text):
    ctypes.windll.user32.MessageBoxW(0, text, title, 0x40 | 0x1)

def check_for_missing_invoices():
    import os
    import openpyxl
    import subprocess
    import pygetwindow as gw
    import fnmatch
    from time import sleep
    import tkinter as tk
    from tkinter import messagebox
    import subprocess
    from pywinauto import Application

    def prompt_to_continue(invoice_number):
        root = tk.Tk()
        root.wm_title("Next Invoice")  # Set the window title

        # Make the window always on top
        root.attributes('-topmost', True)

        # Specify the position of the window (x=200, y=200)
        root.geometry("+400+400")

        tk.Label(root, text=f"Press Enter to view the next missing invoice ({invoice_number}).").grid(row=0)

        def on_button_click(event=None):
            root.destroy()

        # Bind the Enter key to the submit function
        root.bind('<Return>', on_button_click)

        # Update idle tasks, lift and focus the window
        root.update_idletasks()
        root.lift()
        root.focus_force()
        root.wait_window()  # Replace mainloop() with wait_window()

    # Step 1: Read unique invoice numbers from the Excel sheet
    workbook_path = r"C:/Users/Andy Weigl/OneDrive - Kodiak Cakes/Hub Group/Hub Group Upload.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    data_sheet = workbook["Data"]

    unique_invoice_numbers = set()
    for row in range(2, data_sheet.max_row + 1):
        invoice_number = str(data_sheet.cell(row=row, column=1).value)
        unique_invoice_numbers.add(invoice_number)

    # Step 2: Extract invoice numbers from the PDF file names
    pdf_folder_path = "C:/Users/Andy Weigl/OneDrive - Kodiak Cakes/Hub Group/Invoices"
    pdf_files = [f for f in os.listdir(pdf_folder_path) if f.endswith(".pdf")]

    invoice_numbers_from_files = set()
    for pdf_file in pdf_files:
        invoice_number = str(pdf_file.split("_")[0])
        invoice_numbers_from_files.add(invoice_number)

    # Step 3: Compare the two sets of invoice numbers
    missing_invoices = unique_invoice_numbers.symmetric_difference(invoice_numbers_from_files)

    # Check if both sets of invoice numbers are equal and exit early if they are
    if unique_invoice_numbers == invoice_numbers_from_files:
        print("All invoice numbers match. No further action needed.")

        return

    print(invoice_numbers_from_files)
    print(unique_invoice_numbers)

    # Step 4: Notify the user if any invoice numbers are missing
    pdf_viewer = r"C:\Program Files\Nitro\Pro\13\NitroPDF.exe"
    previous_pdf_process = None
    first_iteration = True

    # Load the Excel workbook and select the desired worksheet
    workbook_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Hub Group Upload.xlsx"
    os.startfile(workbook_path)

    if missing_invoices:
        print("The following invoices are missing and will need to be input manually:")
        # You may need to create a hidden tkinter root window before using messagebox
        root = tk.Tk()
        root.withdraw()

        messagebox.showinfo("Information",
                            "Missing invoice(s) detected. Opening the invoice(s) pdf and excel. \nPlease enter manually.")

        # Destroy the hidden root window after the messagebox is closed
        root.destroy()
        for invoice_number in missing_invoices:
            pdf_file_name = f"{invoice_number}_Invoice.pdf"
            pdf_file_path = os.path.join(pdf_folder_path, pdf_file_name)

            if os.path.exists(pdf_file_path):
                print(f" - Invoice {invoice_number}")

                pdf_process = subprocess.Popen([pdf_viewer, pdf_file_path])

                if first_iteration:
                    sleep(3)  # Adjust this sleep time as needed
                    first_iteration = False
                else:
                    sleep(1)

                pdf_window = None
                for win in gw.getAllWindows():
                    if fnmatch.fnmatch(win.title, f"{pdf_file_name} - *"):
                        pdf_window = win
                        break

                if pdf_window is None:
                    print(f"Unable to find the NitroPDF window for invoice number: {invoice_number}")
                    pdf_process.kill()
                    continue
                pdf_window.activate()

                prompt_to_continue(invoice_number)

                if previous_pdf_process is not None:
                    previous_pdf_process.kill()

                previous_pdf_process = pdf_process

            else:
                print(f"PDF file not found for invoice number: {invoice_number}")

        # Close any remaining NitroPDF processes
        if previous_pdf_process is not None:
            previous_pdf_process.kill()

        # Close the Excel file
        app = Application().connect(title_re=".*Hub Group Upload.xlsx.*", class_name="XLMAIN")
        excel_window = app.top_window()
        excel_window.close()

    else:
        print("All invoices have been properly processed.")

# Function to check if both 'Ctrl' and '`' keys are pressed
def check_ctrl_and_backtick():
    return keyboard.is_pressed('ctrl') and keyboard.is_pressed('`')

# Function to minimize a specific window with a given title
def minimize_window_with_title(target_title):
    def callback(hwnd, _):
        window_title = win32gui.GetWindowText(hwnd)
        if target_title in window_title and win32gui.IsWindowVisible(hwnd) and not win32gui.IsIconic(hwnd):
            win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)

    win32gui.EnumWindows(callback, None)

def close_excel_with_temp_file(window_title):
    def window_callback(hwnd, _):
        try:
            window_text = win32gui.GetWindowText(hwnd)
            if window_title.lower() in window_text.lower():
                # Get the window coordinates
                rect = win32gui.GetWindowRect(hwnd)
                x, y, w, _ = rect

                # Move the mouse cursor to the close button and click
                close_button_x = w - 20  # Updated to top right corner
                close_button_y = y + 20
                pyautogui.moveTo(close_button_x, close_button_y, duration=0.1)  # Reduced duration for faster movement
                pyautogui.click()

                # Break the loop
                return False
        except Exception as ex:
            print(f"Error encountered while closing window: {ex}")

    try:
        win32gui.EnumWindows(window_callback, None)
    except Exception as ex:
        print(f"Error encountered during EnumWindows: {ex}")

def bring_temp_file_window_to_front(temp_file_name):
    window_found = False
    while not window_found:
        windows = gw.getWindowsWithTitle(temp_file_name)
        if windows:
            window = windows[0]
            if not window.isActive:
                window.activate()
                time.sleep(0.5)
            window_found = True
        else:
            time.sleep(0.5)

def show_popup(message):
    MB_OK = 0x0
    MB_ICONINFORMATION = 0x40
    MB_SYSTEMMODAL = 0x1000

    ctypes.windll.user32.MessageBoxW(0, message, "Message", MB_OK | MB_ICONINFORMATION | MB_SYSTEMMODAL)

# Get a list of all PDF files in the folder
pdf_files = [f for f in os.listdir(pdf_folder_path) if f.lower().endswith('.pdf')]

# Specify the folder containing the PDF files
pdf_folder_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Invoices"

# Load the existing workbook
excel_file_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Hub Group Upload.xlsx"

while is_excel_file_open(excel_file_path):
    messagebox.showwarning("File Open", "Please close the Hub Group Upload.xlsx file to continue.")

try:
    workbook = openpyxl.load_workbook(excel_file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save(excel_file_path)

# Select the "Data" sheet or create it if it doesn't exist
sheet_name = 'Data'
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    sheet = workbook.create_sheet(sheet_name)

# Clear the "data" sheet and set the headers
sheet.delete_rows(1, sheet.max_row)
sheet.delete_cols(1, sheet.max_column)
headers = ["Invoice Number", "PO Number", "Date", "Charge Type", "Amount", "Location", "Charge Type Modified", "Order Number", "Line Memo", "Location for Line Memo"]
sheet.append(headers)

# Process each PDF file
for pdf_file in pdf_files:
    pdf_file_path = os.path.join(pdf_folder_path, pdf_file)

    print(f"Processing file: {pdf_file}")  # Add this line to print the file name


    # Extract invoice data from the PDF
    invoice_number, reference_numbers, invoice_date, first_page_text = extract_invoice_data_from_pdf(pdf_file_path)

    # Check for "RJW"
    location = check_for_rjw(first_page_text)

    # Extract charges and amounts from the PDF
    charges, first_page_text = extract_charges_from_pdf(pdf_file_path)

    print(charges)
    print(invoice_number)
    print(reference_numbers)  # Updated line
    print(invoice_date)

    num_po_numbers = len(reference_numbers)
    if reference_numbers:
        modified_charges = [(charge, float(amount) / num_po_numbers) for charge, amount in charges]
    else:
        modified_charges = [(charge, float(amount)) for charge, amount in charges]

    rows = []
    if reference_numbers:  # If there are PO numbers
        for idx, reference_number in enumerate(reference_numbers):  # Loop through each PO number
            for charge_idx, (charge, amount) in enumerate(modified_charges):  # Loop through each modified charge
                if idx == num_po_numbers - 1:  # If this is the last PO number
                    # Calculate the amount already distributed for this charge
                    distributed_amount = round(amount, 2) * (num_po_numbers - 1)
                    # Adjust the amount for the last PO number to match the expected total
                    amount = float(charges[charge_idx][1]) - distributed_amount

                row = [invoice_number, reference_number, invoice_date, charge, round(amount, 2), location]
                rows.append(row)
    else:  # If there are no PO numbers
        for charge_idx, (charge, amount) in enumerate(modified_charges):  # Loop through each modified charge
            row = [invoice_number, "", invoice_date, charge, round(amount, 2),
                   location]  # Use an empty string for the missing PO number
            rows.append(row)

    # Add the data rows to the sheet
    for row in rows:
        sheet.append(row)

# Convert columns A and E in the "Data" sheet to numbers
for row_num in range(2, sheet.max_row + 1):
    invoice_num = sheet.cell(row=row_num, column=1).value
    amount = sheet.cell(row=row_num, column=5).value
    order_num = sheet.cell(row=row_num, column=8).value

    if isinstance(invoice_num, str) and invoice_num.isdigit():
        sheet.cell(row=row_num, column=1).value = int(invoice_num)

    if isinstance(amount, str) and re.match(r'^\d+(\.\d{2})?$', amount):
        sheet.cell(row=row_num, column=5).value = float(amount)

    if isinstance(order_num, str) and order_num.isdigit():
        sheet.cell(row=row_num, column=8).value = int(order_num)

# Modify the "Charge Type Modified" column (Column G) based on the values in Column D
for row_num in range(2, sheet.max_row + 1):
    charge_type = sheet.cell(row=row_num, column=4).value
    modified_charge_type = modify_charge_type(charge_type)
    sheet.cell(row=row_num, column=7).value = modified_charge_type

# Get the number of invoices imported
num_invoices = len(pdf_files)

workbook.save(excel_file_path)

# Display a message box with the number of invoices imported and additional instruction
MB_SYSTEMMODAL = 0x1000
message = f"{num_invoices} invoices have been imported into Excel.\nPlease enter Order Numbers."
ctypes.windll.user32.MessageBoxW(0, message, "Import Complete", MB_SYSTEMMODAL)

# Save the workbook before calling the function
workbook.save(excel_file_path)

# Close the workbook
workbook.close()

# Call the function from the second script
view_hub_group_invoices()

# Close the workbook
workbook.close()

# Reload the workbook after the user has manually entered the order numbers
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Update Column I and Column J for the data rows
for row_num in range(2, sheet.max_row + 1):
    col_f_value = sheet.cell(row=row_num, column=6).value
    if col_f_value == "RJW Logistics W10":
        sheet.cell(row=row_num, column=10).value = "RJW"
    else:
        sheet.cell(row=row_num, column=10).value = "NOT RJW, CHECK INVOICE"

    # Update Column I
    col_b_value = sheet.cell(row=row_num, column=2).value
    col_h_value = sheet.cell(row=row_num, column=8).value
    col_j_value = sheet.cell(row=row_num, column=10).value
    col_g_value = sheet.cell(row=row_num, column=7).value

    values_to_concatenate = [col_b_value, col_h_value, col_j_value, col_g_value]
    concatenated_value = "; ".join([str(val) for val in values_to_concatenate[:2] if val is not None])
    concatenated_value += " " + " ".join([str(val) for val in values_to_concatenate[2:] if val is not None])
    sheet.cell(row=row_num, column=9).value = concatenated_value

# Save the workbook before calling the function
workbook.save(excel_file_path)

# Close the workbook
workbook.close()

check_for_missing_invoices()

# Close the workbook
workbook.close()

# Reload the workbook after the user has manually entered the order numbers
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Select the "Upload Template" sheet or create it if it doesn't exist
upload_sheet_name = 'Upload Template'
if upload_sheet_name in workbook.sheetnames:
    upload_sheet = workbook[upload_sheet_name]
else:
    upload_sheet = workbook.create_sheet(upload_sheet_name)

# Clear the "Upload Template" sheet
upload_sheet.delete_rows(1, upload_sheet.max_row)

# Set the headers in the "Upload Template" sheet
upload_headers = [
    "Vendor Internal ID", "Location", "Location Internal ID", "PO", "Customer",
    "Customer Internal ID", "Division", "Department", "Reference #", "Date",
    "Expense", "Expense Internal ID", "Item", "Item Internal ID", "Amount",
    "Line Memo"
]
upload_sheet.append(upload_headers)

# Get a reference to the "Upload Template" sheet
upload_sheet = workbook["Upload Template"]

# Clear the contents of the "Upload Template" sheet, starting from row 2
clear_worksheet_contents(upload_sheet, start_row=2)

# Transfer data from the "Data" sheet to the "Upload Template" sheet
for row_num in range(2, sheet.max_row + 1):
    new_row_num = upload_sheet.max_row + 1
    upload_sheet.cell(row=new_row_num, column=1).value = 5905
    upload_sheet.cell(row=new_row_num, column=2).value = sheet.cell(row=row_num, column=6).value
    upload_sheet.cell(row=new_row_num, column=3).value = '=VLOOKUP(B{},Locations!B:C,2,FALSE)'.format(new_row_num)
    upload_sheet.cell(row=new_row_num, column=4).value = sheet.cell(row=row_num, column=2).value
    upload_sheet.cell(row=new_row_num, column=8).value = 'Operations : Manufacturing'
    upload_sheet.cell(row=new_row_num, column=9).value = sheet.cell(row=row_num, column=1).value
    upload_sheet.cell(row=new_row_num, column=10).value = sheet.cell(row=row_num, column=3).value
    upload_sheet.cell(row=new_row_num, column=13).value = '=IF(B{}="RJW Logistics W10","Landed Cost - RJW - Freight","NOT RJW")'.format(new_row_num)
    upload_sheet.cell(row=new_row_num, column=14).value = '=VLOOKUP(M{},Items!A:C,3,FALSE)'.format(new_row_num)
    upload_sheet.cell(row=new_row_num, column=15).value = sheet.cell(row=row_num, column=5).value
    upload_sheet.cell(row=new_row_num, column=16).value = sheet.cell(row=row_num, column=9).value

# Save the workbook before calling the function
workbook.save(excel_file_path)

# Close the workbook
workbook.close()

exe_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Python Scripts\Python Venv - Desktop\dist\Get Division.exe"

# Start the subprocess using Popen
process = subprocess.Popen([exe_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

# Wait for the subprocess to finish
stdout, stderr = process.communicate()

# Minimize the specific window
target_title = "NetSuite (Kodiak Cakes LLC) — Mozilla Firefox"
minimize_window_with_title(target_title)

# Close the workbook
workbook.close()

# Reload the workbook after the user has manually entered the order numbers
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

os.startfile(excel_file_path)

message = "Please review the data in the worksheet, then save and close the workbook. \nPress Control + ` to continue the script."
show_popup(message)

print(f"Press control + ` to continue")
wait_for_key_combination()

# Close the workbook
workbook.close()

# Reload the workbook after the user has manually entered the order numbers
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Get a reference to the "Upload Template" sheet
upload_sheet = workbook["Upload Template"]

# Save the workbook before calling the function
workbook.save(excel_file_path)

workbook_path = r"C:/Users/Andy Weigl/OneDrive - Kodiak Cakes/Hub Group/Hub Group Upload.xlsx"

# Reload the workbook after user has manually entered the order numbers
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

# Clear the contents of the "Hub Group Primary" sheet
hub_group_primary_sheet_name = 'Hub Group Primary'
if hub_group_primary_sheet_name in workbook.sheetnames:
    hub_group_primary_sheet = workbook[hub_group_primary_sheet_name]
    clear_sheet(hub_group_primary_sheet)

# Clear the contents of the "Hub Group Expenses" sheet
hub_group_expenses_sheet_name = 'Hub Group Expenses'
if hub_group_expenses_sheet_name in workbook.sheetnames:
    hub_group_expenses_sheet = workbook[hub_group_expenses_sheet_name]
    clear_sheet(hub_group_expenses_sheet)

# Clear the contents of the "Hub Group Items" sheet
hub_group_items_sheet_name = 'Hub Group Items'
if hub_group_items_sheet_name in workbook.sheetnames:
    hub_group_items_sheet = workbook[hub_group_items_sheet_name]
    clear_sheet(hub_group_items_sheet)

# Get the "Upload Template" worksheet
upload_template_sheet_name = 'Upload Template'
upload_template_sheet = workbook[upload_template_sheet_name]

# Copy the specified columns from the "Upload Template" worksheet to the "Hub Group Primary" worksheet
copy_column(upload_template_sheet, 1, hub_group_primary_sheet, 1)  # Copy Column A to Column A
copy_column(upload_template_sheet, 10, hub_group_primary_sheet, 2)  # Copy Column J to Column B
copy_column(upload_template_sheet, 9, hub_group_primary_sheet, 3)  # Copy Column I to Column C
copy_column(upload_template_sheet, 16, hub_group_primary_sheet, 4)  # Copy Column P to Column D

# Remove duplicates based on Column C in the "Hub Group Primary" worksheet
remove_duplicates(hub_group_primary_sheet, 2, 1)

# Replace these with your actual sheet names/objects
upload_template_sheet = workbook["Upload Template"]
hub_group_items_sheet = workbook["Hub Group Items"]

# Copy columns A-P (1-16)
copy_columns(upload_template_sheet, hub_group_items_sheet, start_col=1, end_col=16)

workbook.save(excel_file_path)

# Load the source workbook
source_workbook = openpyxl.load_workbook(r'C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Hub Group Upload.xlsx')

# Create date-based folder and file name
current_date = datetime.now().strftime("%m.%d")
csv_upload_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\CSV Upload"
date_folder_path = os.path.join(csv_upload_path, current_date)

if not os.path.exists(date_folder_path):
    os.makedirs(date_folder_path)

# Process both worksheets
process_worksheet(source_workbook, "Hub Group Primary", date_folder_path)

# Load the Excel workbook and select the desired worksheet
file_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\Hub Group Upload.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook["Hub Group Items"]

# Activate the "Hub Group Items" worksheet
workbook.active = worksheet

# Save the workbook to a temporary file (so that the worksheet is activated when opened)
temp_file = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Excel Macro Workbooks\temp_file.xlsx"
workbook.save(temp_file)

# Open the temporary file with the default application (Excel)
os.startfile(temp_file)

# Give Excel some time to open
time.sleep(3)

# Bring the temp file window to the front
temp_file_name = os.path.basename(temp_file)
bring_temp_file_window_to_front(temp_file_name)

# Send the specified keystrokes
pyautogui.hotkey("ctrl", "a")
time.sleep(0.1)
pyautogui.hotkey("ctrl", "c")
time.sleep(0.1)
pyautogui.hotkey("ctrl", "alt", "v")
time.sleep(0.1)
pyautogui.press("v")
time.sleep(0.1)
pyautogui.press("enter")
time.sleep(0.1)
pyautogui.press("escape")
time.sleep(0.1)
pyautogui.press("escape")
time.sleep(0.1)

window_title = "temp_file.xlsx  -  Group - Excel"
close_excel_with_temp_file(window_title)

# Give Excel some time to close
time.sleep(2)

# Load the updated data from the temporary Excel file
workbook = openpyxl.load_workbook(temp_file, data_only=True)
worksheet = workbook["Hub Group Items"]

# Get the current date in mm.dd format
current_date = datetime.now().strftime("%m.%d")

# Set the folder path where the new folder will be created
base_folder_path = r"C:\Users\Andy Weigl\OneDrive - Kodiak Cakes\Hub Group\CSV Upload"

# Create the new folder using the current date
new_folder_path = os.path.join(base_folder_path, current_date)
os.makedirs(new_folder_path, exist_ok=True)

# Set the output CSV file name using the current date
output_csv = f"Items - {current_date}.csv"

# Set the full path of the output CSV file
output_csv_path = os.path.join(new_folder_path, output_csv)

# Save the entire worksheet to the output CSV file
with open(output_csv_path, mode="w", newline="") as csv_file:
    csv_writer = csv.writer(csv_file)

    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows():
        # Extract the values of the cells in the current row
        row_values = [cell.value for cell in row]
        # Write the row values to the CSV file
        csv_writer.writerow(row_values)

# Delete the temporary file
os.remove(temp_file)

# Display a message box with the number of invoices imported and additional instruction
MB_SYSTEMMODAL = 0x1000
message = f"The CSV files have been saved and are ready to be uploaded to NetSuite."
ctypes.windll.user32.MessageBoxW(0, message, "Script Complete", MB_SYSTEMMODAL)

# Save the workbook
workbook.save(excel_file_path)
workbook.close()